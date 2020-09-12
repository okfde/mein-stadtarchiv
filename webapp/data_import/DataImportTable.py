# encoding: utf-8

"""
Copyright (c) 2017, Ernesto Ruge
All rights reserved.
Redistribution and use in source and binary forms, with or without modification, are permitted provided that the following conditions are met:
1. Redistributions of source code must retain the above copyright notice, this list of conditions and the following disclaimer.
2. Redistributions in binary form must reproduce the above copyright notice, this list of conditions and the following disclaimer in the documentation and/or other materials provided with the distribution.
3. Neither the name of the copyright holder nor the names of its contributors may be used to endorse or promote products derived from this software without specific prior written permission.
THIS SOFTWARE IS PROVIDED BY THE COPYRIGHT HOLDERS AND CONTRIBUTORS "AS IS" AND ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE ARE DISCLAIMED. IN NO EVENT SHALL THE COPYRIGHT HOLDER OR CONTRIBUTORS BE LIABLE FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION) HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF SUCH DAMAGE.
"""

import os
import csv
from typing import Optional
from openpyxl import Workbook, load_workbook
from flask import current_app
from ..models import Category, Document



class DataImportTable:
    def __init__(self, filename: str, csv_autodetect: Optional[bool] = True, mapping: Optional[list] = None, category: Category = None):
        self.filename = filename
        self.csv_autodetect = csv_autodetect
        self.path = os.path.join(current_app.config['TEMP_UPLOAD_DIR'], self.filename)
        self.filetype = self.path.split('.')[1]
        self.data = []
        self.load()

    def load(self):
        """
        if self.filetype == 'xls':
            self.workbook = Workbook()
            xls_workbook = open_workbook(filename=self.path)
            for i in range(0, xls_workbook.nsheets):
                xls_sheet = xls_workbook.sheet_by_index(i)
                sheet = self.workbook.active if i == 0 else self.workbook.create_sheet()
                sheet.title = xls_sheet.name
                for row in range(0, xls_sheet.nrows):
                    for col in range(0, xls_sheet.ncols):
                        sheet.cell(row=row + 1, column=col + 1).value = xls_sheet.cell_value(row, col)
        """
        if self.filetype == 'xlsx':
            self.workbook = load_workbook(self.path)
            self.sheet = self.workbook.active
            for row in self.sheet.iter_rows():
                line = []
                for field in row:
                    line.append(field.value)
                self.data.append(line)

        if self.filetype == 'csv':
            self.workbook = Workbook()
            self.sheet = self.workbook.active
            with open(self.path) as csvfile:
                dialect = csv.Sniffer().sniff(csvfile.read(1024))
                csvfile.seek(0)
                csv_reader = csv.reader(csvfile, dialect=dialect)
                for csv_row in csv_reader:
                    self.data.append(csv_row)

    @property
    def header(self):
        if not len(self.data):
            return []
        return self.data[0]

    @property
    def preview(self):
        return self.data[1:11]

    def save(self, category, mapping):
        line_number = 1
        for line in self.data:
            if line_number == 1:
                line_number += 1
                continue
            if 'identifier' in mapping:
                uid = line[mapping.index('identifier')]
            else:
                uid = line_number
            document = Document.objects(category=category, uid=uid).first()
            if not document:
                document = Document()
                document.uid = uid
            for i in range(0, len(mapping)):
                setattr(document, mapping[i], line[i])
            document.save()
